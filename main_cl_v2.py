# -*- coding: utf-8 -*-
"""
Created on Sat Mar 14 00:31:36 2020

@author: Juan Renatto
"""
from selenium import webdriver
import pandas as pd
from bs4 import BeautifulSoup
import requests
import time
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from re import compile
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date, timedelta
import smtplib
import itertools
import openpyxl
import statistics
import logging
import traceback
from tabulate import tabulate
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import pyautogui
#SELENIUM ---------------------------XXXXXXXXXXXXXXXXX

#import logging

#logger=logging.getLogger()

urlply= "https://www.flashs\core.pe/futbol/peru/liga-1/resultados/"
executable_path='C:\Program Files (x86)\chromedriver.exe'
liemp = ['0','0','0','0','0:0','0','NULL']
today = date.today()
patha = 'C:/Users/mmo_2/OneDrive/oc/data/'
pathb = 'C:/Users/mmo_2/OneDrive/oc/resultados/'
pathc = 'C:/Users/mmo_2/OneDrive/oc/chose/'
pathd = 'C:/Users/mmo_2/OneDrive/oc/'
pyautogui.PAUSE = 2.5
pyautogui.FAILSAFE = False
file_name = 'data_' + str(today) + '.xlsx' #data_2023-02-18
#file_name = 'data_' + '2023-02-17' + '.xlsx'
file_nameb = 'res_' + str(today) + '.xlsx'
file_namec = 'datc_' + str(today) + '.xlsx' #data_2023-02-18
file_named = 'betdash.xlsx' #data_2023-02-18


def resul_cuo(sheet, result_col):
    conteo = 2
    for row in range(conteo, sheet.max_row + 1):
            result =  '=IF(ISNA(VLOOKUP(DN'+str(conteo)+',HIST_EFIC!G:H,2,FALSE)),"",VLOOKUP(DN'+str(conteo)+',HIST_EFIC!G:H,2,FALSE))'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1

def resul_log(sheet, result_col):
    conteo = 2
    for row in range(conteo, sheet.max_row + 1):
            result =  '=IF(ISNA(VLOOKUP(DN'+str(conteo)+',HIST_CUOT!G:H,2,FALSE)),"",VLOOKUP(DN'+str(conteo)+',HIST_CUOT!G:H,2,FALSE))'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1



def kort3(sheet, result_col):
    conteo = 2
    for row in range(conteo, sheet.max_row + 1):
            result =  '=FIND("|",CY'+str(conteo)+',FIND("|",CY'+str(conteo)+',FIND("|",CY'+str(conteo)+',1)+1)+1)'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1



def kort4(sheet, result_col):
    conteo = 2
    for row in range(conteo, sheet.max_row + 1):
            result =  '=FIND("|",CY'+str(conteo)+',DJ'+str(conteo)+'+1)'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1



def keylog(sheet, result_col):
    conteo = 2
    for row in range(conteo, sheet.max_row + 1):
            result =  '=MID(CY'+str(conteo)+',DJ'+str(conteo)+'+1,DK'+str(conteo)+'-DJ'+str(conteo)+'-1)'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1


def kort5(sheet, result_col):
    conteo = 2
    for row in range(conteo, sheet.max_row + 1):
            result =  '=FIND("-",DL'+str(conteo)+',1)'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1


def keycuo(sheet, result_col):
    conteo = 2
    for row in range(conteo, sheet.max_row + 1):
            result =  '=MID(DL'+str(conteo)+',1,DM'+str(conteo)+'-2)'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1


# Define a function to filter a column based on the values being less than or greater than a threshold
def filter_column(sheet, filt1, col2, filt2, col4,filt3, col6, threshold_1 , threshold_2, threshold_3):
    filtered_values = []
    for row in range(2, sheet.max_row + 1):
        desve1 = sheet.cell(row=row, column=filt1)
        averg1 = sheet.cell(row=row, column=filt2)
        opinio = sheet.cell(row=row, column=filt3)
        if desve1.value is not None and averg1.value is not None and opinio.value is not None:
            if desve1.value <= threshold_1 and averg1.value >= threshold_2 and opinio.value == threshold_3:
                part = sheet.cell(row=row, column=col2)
                expe = sheet.cell(row=row, column=col4)
                opin = sheet.cell(row=row, column=col6)
                filtered_values.append(part.value)
                filtered_values.append(expe.value)
                filtered_values.append(opin.value)
    return filtered_values

def div_columns(sheet, col1, col2, result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        try :
            result = float(cell1.value) / float(cell2.value)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def same(sheet, col1, result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        try:
            result = cell1.value
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def div_var_columns(sheet, col1, col2, col3, result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        try :
            result = (float(cell1.value) - float(cell2.value)) / float(cell3.value)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0


def sum_tot (sheet, col1, col2, col3, col4, col5, col6, col7, col8, col9, result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        cell4 = sheet.cell(row=row, column=col4)
        cell5 = sheet.cell(row=row, column=col5)
        cell6 = sheet.cell(row=row, column=col6)
        cell7 = sheet.cell(row=row, column=col7)
        cell8 = sheet.cell(row=row, column=col8)
        cell9 = sheet.cell(row=row, column=col9)
        try:
            result = (float(cell1.value) + float(cell2.value) +float(cell3.value) +float(cell4.value)+float(cell5.value)+float(cell6.value)+float(cell7.value)+float(cell8.value)+float(cell9.value))
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def sum_tot_abs (sheet, col1, col2, col3, col4, col5, col6, col7, col8, col9, result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        cell4 = sheet.cell(row=row, column=col4)
        cell5 = sheet.cell(row=row, column=col5)
        cell6 = sheet.cell(row=row, column=col6)
        cell7 = sheet.cell(row=row, column=col7)
        cell8 = sheet.cell(row=row, column=col8)
        cell9 = sheet.cell(row=row, column=col9)
        try:
            result = (float(cell1.value) + float(cell2.value) +float(cell3.value) +float(cell4.value)+float(cell5.value)+float(cell6.value)+float(cell7.value)+float(cell8.value)+float(cell9.value))
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = abs(result)
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def desvest (sheet, col1, col2, col3, col4, col5, col6, col7, col8, col9, result_col):
    for row in range(2, sheet.max_row + 1):
        numbers=[]
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        cell4 = sheet.cell(row=row, column=col4)
        cell5 = sheet.cell(row=row, column=col5)
        cell6 = sheet.cell(row=row, column=col6)
        cell7 = sheet.cell(row=row, column=col7)
        cell8 = sheet.cell(row=row, column=col8)
        cell9 = sheet.cell(row=row, column=col9)
        numbers.append(float(cell1.value))
        numbers.append(float(cell2.value))
        numbers.append(float(cell3.value))
        numbers.append(float(cell4.value))
        numbers.append(float(cell5.value))
        numbers.append(float(cell6.value))
        numbers.append(float(cell7.value))
        numbers.append(float(cell8.value))
        numbers.append(float(cell9.value))
        try:
            result = statistics.stdev(numbers)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:

            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0


def div_var_columns_50(sheet, col1, col2, col3, result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        try:
            result = (float(cell1.value) - float(cell2.value)) / float(cell3.value)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result * 0.5
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0


def div_gc_columns(sheet, col1, col2,  result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        try:
            result = (float(cell1.value[cell1.value.find(":")+1:]) / float(cell2.value))
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0


def concatcuo(sheet, col1, col2,  result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        try:
            result = str(cell1.value) + ":" + str(cell2.value)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 'NULL'

def concatkey(sheet, col1, col2,  result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        try:
            result = str(cell1.value) + " - " + str(cell2.value)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 'NULL'



def valida(sheet, col1, col2, col3, result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        if cell1.value == "NULL - NULL" or int(cell2.value) < 7:
            result = "NO"
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        else:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 'SI'

def div_gf_columns(sheet, col1, col2,  result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        try:
            result = (float(cell1.value[:cell1.value.find(":")]) / float(cell2.value))
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def resultado(sheet, col1, col2, col3, result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)

        if cell1.value <= -4.5 and cell2.value <= -0.5:
            result = "VISITA-EMPATE"
        elif cell1.value <= -3.5 and cell2.value <= -0.4:
            result = "VISITA-EMPATE"
        elif cell1.value >= 3.5 and cell2.value >= 0.4:
            result = "LOCAL"
        elif cell1.value >= 0 and cell1.value < 3.5:
            result = "LOCAL-EMPATE"
        elif cell1.value > -3.5 and cell1.value < 0:
            result = "LOCAL-EMPATE"
        else:
            result = "EMPATE"
        result_cell = sheet.cell(row=row, column=result_col)
        result_cell.value = result

        #if cell1.value <= -4.5 and cell2.value <= -0.5:
        #    result = "VISITA-EMPATE"
        #elif cell1.value <= -3.5 and cell2.value <= -0.4:
        #    result = "VISITA-EMPATE"
        #elif cell1.value >= 3.5 and cell2.value >= 0.4:
        #    result = "LOCAL"
        #elif cell1.value >= 0 and cell1.value < 3.5:
        #    result = "LOCAL-EMPATE"
        #elif cell1.value > -3.5 and cell1.value < 0:
        #    result = "LOCAL-EMPATE"
        #else:
        #    result = "EMPATE"
        #result_cell = sheet.cell(row=row, column=result_col)
        #result_cell.value = result

def average_dist(sheet, col1, col2, col3,  result_col):
    for row in range(2, sheet.max_row + 1):
        numbers = []
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        numbers.append(float(cell1.value))
        numbers.append(float(cell2.value))
        numbers.append(float(cell3.value))
        try:
            non_zero_values = [value for value in numbers if value != 0]
            if non_zero_values:
                result = sum(non_zero_values) / len(non_zero_values)
                result_cell = sheet.cell(row=row, column=result_col)
                result_cell.value = result
            else:
                result_cell = sheet.cell(row=row, column=result_col)
                result_cell.value = 0
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0


def average_dist(sheet, col1, col2, col3,  result_col):
    for row in range(2, sheet.max_row + 1):
        numbers = []
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        numbers.append(float(cell1.value))
        numbers.append(float(cell2.value))
        numbers.append(float(cell3.value))
        try:
            non_zero_values = [value for value in numbers if value != 0]
            if non_zero_values:
                result = sum(non_zero_values) / len(non_zero_values)
                result_cell = sheet.cell(row=row, column=result_col)
                result_cell.value = result
            else:
                result_cell = sheet.cell(row=row, column=result_col)
                result_cell.value = 0
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0


def average_dist_abs(sheet, col1, col2, col3,  result_col):
    for row in range(2, sheet.max_row + 1):
        numbers = []
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        numbers.append(float(cell1.value))
        numbers.append(float(cell2.value))
        numbers.append(float(cell3.value))
        try:
            non_zero_values = [value for value in numbers if value != 0]
            if non_zero_values:
                result = sum(non_zero_values) / len(non_zero_values)
                result_cell = sheet.cell(row=row, column=result_col)
                result_cell.value = abs(result)
            else:
                result_cell = sheet.cell(row=row, column=result_col)
                result_cell.value = 0
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def find_substring_res(string, substring, start):
    return int(string.find(substring, start) + 1)

def concatkeynovo(sheet, col1, string,  result_col):
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)

        try:
            result = cell1.value[(find_substring_res(cell1.value, string , 1)+1):(find_substring_res(cell1.value, string, find_substring_res(cell1.value, string, 1))-2)]
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            #print(result_cell)
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 'NULL'
            #print(result_cell)




def send_email_y(subject, message, to_email, from_email, password, file_path, workbname):
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(from_email, password)

        # Create a multipart message
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject

        # Add the message body
        body = MIMEText(message)
        msg.attach(body)

        # Add the Excel file as an attachment
        with open(file_path, 'rb') as f:
            attachment = MIMEApplication(f.read(), _subtype='xlsx')
            attachment.add_header('Content-Disposition', 'attachment', filename=workbname)
            msg.attach(attachment)

        # Send the email
        server.sendmail(from_email, to_email, msg.as_string())
        print("Email sent!")
    except Exception as e:
        print(f"Something went wrong... {e}")
    finally:
        server.quit()

def copy_column(src_sheet, dst_sheet, src_col, dst_col):
    #acron = ''
    fila = 0
    for row in src_sheet.iter_rows(values_only=True):
        #fila = len(dst_sheet[dst_col]) + 1
        #print('source max')
        #print(src_sheet.max_row)
        #print('destino max')
        #print(dst_sheet.max_row)
        max_src = src_sheet.max_row
        if max_src != fila:
            fila = fila+1
        else:
            fila = 0
        #print('fila')
        #print(fila)


        #dst_sheet.cell(row=1, column=dst_col).value = row[src_col - 1]
        #dst_sheet.cell(row=dst_sheet.max_row + 1, column=dst_col).value = row[src_col - 1]
        dst_sheet.cell(row=fila, column=dst_col).value = row[src_col - 1]

def send_email_x(subject, message, to_email, from_email, password):
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(from_email, password)
        message = f"Subject: {subject}\n\n{message}"
        server.sendmail(from_email, to_email, message)
        print("Email sent!")
    except Exception as e:
        print(f"Something went wrong... {e}")
    finally:
        server.quit()

subject = "Elecciones del dia."
subjecx = "Todos los pronosticos del dia."
message = "Selecciones del dia. Leyenda: DEV1 = Mayor valor (-), TOV2 = Mayor valor (+), AVV3 = Mayor valor (+), EFIC_1 = Ratio Historico de Exito por Liga segun Logica (+), EFIC_2 = Ratio Historico de Exito por Liga segun Cuotas (+)"
to_email = "juanre_68@hotmail.com"
from_email = "juanre060894@gmail.com"
password = "jdlwjzmbahaxvbwq"





def has_id(tag):
    return tag.has_attr('id')

options = Options()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_argument('--ignore-certificate-errors-spki-list')
options.add_argument('--headless')

segs = random.randrange(6,8)
segc = random.randrange(7,9)



df = pd.DataFrame(
    columns=
            [
                'PARTIDO',
                'PJ_ACUM_1',
                'G_ACUM_1',
                'E_ACUM_1',
                'P_ACUM_1',
                'DIF_ACUM_1',
                'PTS_ACUM_1',
                'EQUIPO_ACUM_1',
                'PJ_ACUM_2',
                'G_ACUM_2',
                'E_ACUM_2',
                'P_ACUM_2',
                'DIF_ACUM_2',
                'PTS_ACUM_2',
                'EQUIPO_ACUM_2',
                'PJ_GENERAL_5_1',
                'G_GENERAL_5_1',
                'E_GENERAL_5_1',
                'P_GENERAL_5_1',
                'DIF_GENERAL_5_1',
                'PTS_GENERAL_5_1',
                'EQUIPO_GENERAL_5_1',
                'PJ_GENERAL_5_2',
                'G_GENERAL_5_2',
                'E_GENERAL_5_2',
                'P_GENERAL_5_2',
                'DIF_GENERAL_5_2',
                'PTS_GENERAL_5_2',
                'EQUIPO_GENERAL_5_2',
                'PJ_LOCAL_5_1',
                'G_LOCAL_5_1',
                'E_LOCAL_5_1',
                'P_LOCAL_5_1',
                'DIF_LOCAL_5_1',
                'PTS_LOCAL_5_1',
                'EQUIPO_LOCAL_5_1',
                'PJ_LOCAL_5_2',
                'G_LOCAL_5_2',
                'E_LOCAL_5_2',
                'P_LOCAL_5_2',
                'DIF_LOCAL_5_2',
                'PTS_LOCAL_5_2',
                'EQUIPO_LOCAL_5_2',
                'PJ_VISITA_5_1',
                'G_VISITA_5_1',
                'E_VISITA_5_1',
                'P_VISITA_5_1',
                'DIF_VISITA_5_1',
                'PTS_VISITA_5_1',
                'EQUIPO_VISITA_5_1',
                'PJ_VISITA_5_2',
                'G_VISITA_5_2',
                'E_VISITA_5_2',
                'P_VISITA_5_2',
                'DIF_VISITA_5_2',
                'PTS_VISITA_5_2',
                'EQUIPO_VISITA_5_2',




             ]
              )

try:


    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
    driver.get("https://www.flashscore.pe/?rd=mismarcadores.com")

    time.sleep(segs)

    number_li_elems=len(WebDriverWait(driver,30).until(EC.visibility_of_all_elements_located((By.XPATH, "//span[@title='Mostrar todos los partidos de esta liga']"))))
    print(number_li_elems)
    for x in range(number_li_elems):
        time.sleep(segc)
        # you have to get the element by index every time, otherwise you will get StaleElement Exception
        #button = driver.find_element(By.XPATH, "//span[@title='Ocultar todos los partidos de esta liga']")
        button = driver.find_element(By.XPATH, "//span[@title='Mostrar todos los partidos de esta liga']")
        #Mostrar todos los partidos de esta liga
        print(button)
        driver.execute_script("arguments[0].click();", button)


    time.sleep(segs)
    html = driver.page_source
    soupply= BeautifulSoup (html,'html.parser')
    tableply = soupply.find('div',class_ ='sportName soccer') #, id='stats_standard'
    #print(tableply)
    tableplx = tableply.find_all(has_id)
    #tableplx = tableply.find_all('div',class_ ='event__match event__match--twoLine')

    print('CORTE')


    idprin = []


    for idx in tableplx:
        idclea = idx.get('id')
        idprin.append(idclea[4:len(idclea)])
    print(idprin)
    print('ingreso cuotas')
    #
    #--------------------------VAMOS
    time.sleep(segc)

    buttob = driver.find_element(By.XPATH,"//div[text()='Cuotas']")

    driver.execute_script("arguments[0].click();", buttob)

    time.sleep(segc)
    time.sleep(segc)
    time.sleep(segc)
    htmlb = driver.page_source

    soupplyb= BeautifulSoup (htmlb,'html.parser')

    driver.quit()



    idpric = []
    cuotas = []
    #idprin = ['zoD4r1Ur','xAXTl3iF'] #test
    #for idc in idprin:
    for idc in idprin:
        idcx= 'g_1_' + idc
        tableplb = soupplyb.find('div',id =idcx) #, id='stats_standard'
        print(tableplb)
        #tablepbx = tableplb.find('div', class_=re.compile('^odds__odd event__odd--odd1'))
        tablepbx=tableplb.find("div", class_=lambda value: value and value.startswith("odds__odd event__odd--odd1"))
        print(tablepbx)
        print(tablepbx.text)
        if tablepbx.text != '-':
            idpric.append(idc)
            cuotas.append(tablepbx.text)
        else:
            print('sin cuota')


    print('CORTE CUOTA')
    print(idpric)
    print(cuotas)
    print('CORTE CUOTA')
    #idpric = ['dWRnrhz0','WYBmFdEU']
    #cuotas = [2.4,2.3]
    #--------------------------VAMOS

    #idpric = ['vDEAKFbc','SxB6Leqi','zqOKchip','baxzyYil','zaM4gZ0G']
    #cuotas = [2.14,2.15,2,2,2]
    #idpric = ['vDEAKFbc','SxB6Leqi']
    #cuotas = [2.14,2.15]
    conteo = 0
    totalp = len(idpric)
    #for key in idpric:

    driverc = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
    for (key,tip) in zip(idpric,cuotas):
        start_time = time.time()
        print(start_time)



        resume = 'https://www.flashscore.pe/partido/'+key+'/#/resumen-del-partido/resumen-del-partido'
        clasif = 'https://www.flashscore.pe/partido/'+key+'/#/clasificacion/table/overall'
        formag = 'https://www.flashscore.pe/partido/'+key+'/#/clasificacion/form/overall/5'
        formal = 'https://www.flashscore.pe/partido/'+key+'/#/clasificacion/form/home/5'
        formav = 'https://www.flashscore.pe/partido/'+key+'/#/clasificacion/form/away/5'
        conteo = conteo + 1
        print(str(conteo)+'/'+str(totalp))

        #BS4 ---------------------------XXXXXXXXXXXXXXXXX 1 RECUPERACION DE EQUIPOS
        dagen = []
        #time.sleep(segs)
        #driverc = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
        print('ingreso')



        driverc.get(resume)
        time.sleep(segs)
        sourcr = driverc.page_source
        #sourcr=requests.get(resume,timeout=10 ).text
        soupr= BeautifulSoup (sourcr,'html.parser')
        #table = soupr.head.title.string
        try:
            table = soupr.head.title.string
        except:
            table = 'NUL 0-0 NUL | Null - Null | Resumen'

        meta_tag = soupr.find('meta', attrs={'name': 'og:description'})
        divis = meta_tag['content']
        print(len(table.split('|')))
        print('1ER NIVEL')
        print(table)
        print(divis)
        print('1ER NIVEL')
        finish_time_1 = time.time()
        print(finish_time_1-start_time)
        dagen.append(table)

        #BS4 ---------------------------XXXXXXXXXXXXXXXXX 1 RECUPERACION DE EQUIPOS

        #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX 2 RECUPERACION DE PUNTOS TOTALES LOCAL Y VISITA
        #driverc.quit()
        driverc.get(clasif)
        time.sleep(segs)
        htmlc = driverc.page_source
        soupplc= BeautifulSoup (htmlc,'html.parser')
        tableplc = soupplc.find_all('div',class_ ='ui-table__row table__row--selected') #, id='stats_standard'
        #tableplx = tableply.find_all('div',class_ ='event__match event__match--twoLine')

        if len(tableplc)==2:
            cquip1 = []
            cquip2 = []
            copqui = tableplc[0].find('a',class_ ='tableCellParticipant__name').string
            cosqui = tableplc[1].find('a',class_ ='tableCellParticipant__name').string
            ctpqui = tableplc[0].find_all('span')
            ctsqui = tableplc[1].find_all('span')

            for datos in ctpqui:
                cquip1.append(datos.string)

            for datos in ctsqui:
                cquip2.append(datos.string)

            cquip1.append(copqui)
            cquip2.append(cosqui)

            print('2DO NIVEL')
            print(cquip1)
            print('-----------------------------')
            print(cquip2)
            print('-----------------------------')
            print('2DO NIVEL')
            finish_time_2 = time.time()
            print(finish_time_2-finish_time_1)

        else:
            finish_time_2 = time.time()
            print(finish_time_2-finish_time_1)
            cquip1=liemp
            cquip2=liemp

        #driverc.quit()
        #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX 2 RECUPERACION DE PUNTOS TOTALES LOCAL Y VISITA



        #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX 3 FORMA 5 GENERAL
        #driverc.quit()
        #time.sleep(segs)

        #driverg = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
        driverc.get(formag)

        time.sleep(segs)
        htmlg = driverc.page_source
        soupplg= BeautifulSoup (htmlg,'html.parser')
        tableplg = soupplg.find_all('div',class_ ='ui-table__row table__row--selected') #, id='stats_standard'
        #tableplx = tableply.find_all('div',class_ ='event__match event__match--twoLine')
        if len(tableplg)==2:
            gquip1 = []
            gquip2 = []
            gopqui = tableplg[0].find('a',class_ ='tableCellParticipant__name').string
            gosqui = tableplg[1].find('a',class_ ='tableCellParticipant__name').string
            gtpqui = tableplg[0].find_all('span')
            gtsqui = tableplg[1].find_all('span')

            for datos in gtpqui:
                gquip1.append(datos.string)

            for datos in gtsqui:
                gquip2.append(datos.string)

            gquip1.append(gopqui)
            gquip2.append(gosqui)

            print('3ER NIVEL')
            print(gquip1)
            print('-----------------------------')
            print(gquip2)
            print('-----------------------------')
            print('3ER NIVEL')
            finish_time_3 = time.time()
            print(finish_time_3-finish_time_2)

            #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX 3 FORMA 5 GENERAL

            #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX 4 FORMA 5 lOCAL
            #driverc.quit()
            #time.sleep(segs)


            driverc.get(formal)

            time.sleep(segs)
            htmll = driverc.page_source
            souppll= BeautifulSoup (htmll,'html.parser')
            tablepll = souppll.find_all('div',class_ ='ui-table__row table__row--selected') #, id='stats_standard'
            #tableplx = tableply.find_all('div',class_ ='event__match event__match--twoLine')
            if len(tablepll)==2:

                lquip1 = []
                lquip2 = []
                lopqui = tablepll[0].find('a',class_ ='tableCellParticipant__name').string
                losqui = tablepll[1].find('a',class_ ='tableCellParticipant__name').string
                ltpqui = tablepll[0].find_all('span')
                ltsqui = tablepll[1].find_all('span')

                for datos in ltpqui:
                    lquip1.append(datos.string)

                for datos in ltsqui:
                    lquip2.append(datos.string)

                lquip1.append(lopqui)
                lquip2.append(losqui)

                print('4TO NIVEL')
                print(lquip1)
                print('-----------------------------')
                print(lquip2)
                print('-----------------------------')
                print('4TO NIVEL')
                finish_time_4 = time.time()
                print(finish_time_4-finish_time_3)
            else:
                lquip1=liemp
                lquip2=liemp
                finish_time_4 = time.time()
                print(finish_time_4-finish_time_3)


            #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX 4 FORMA 5 LOCAL



            #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX 5 FORMA 5 VISITANTE
            #driverc.quit()
            #time.sleep(segs)


            driverc.get(formav)

            time.sleep(segs)
            htmlv = driverc.page_source
            soupplv= BeautifulSoup (htmlv,'html.parser')
            tableplv = soupplv.find_all('div',class_ ='ui-table__row table__row--selected') #, id='stats_standard'
            #tableplx = tableply.find_all('div',class_ ='event__match event__match--twoLine')
            if len(tableplv)==2:
                vquip1 = []
                vquip2 = []
                vopqui = tableplv[0].find('a',class_ ='tableCellParticipant__name').string
                vosqui = tableplv[1].find('a',class_ ='tableCellParticipant__name').string
                vtpqui = tableplv[0].find_all('span')
                vtsqui = tableplv[1].find_all('span')

                for datos in vtpqui:
                    vquip1.append(datos.string)

                for datos in vtsqui:
                    vquip2.append(datos.string)

                vquip1.append(vopqui)
                vquip2.append(vosqui)

                #print(tableplc)
                #print('-----------------------------')
                #print(nopqui)
                #print('-----------------------------')
                #print(nosqui)
                #print('-----------------------------')
                print('5TO NIVEL')
                print(vquip1)
                print('-----------------------------')
                print(vquip2)
                print('-----------------------------')
                print('5TO NIVEL')
                finish_time_5 = time.time()
                print(finish_time_5-finish_time_4)
            else:
                vquip1=liemp
                vquip2=liemp
                finish_time_5 = time.time()
                print(finish_time_5-finish_time_4)
            #driverc.quit()

        else:
            #driverc.quit()
            gquip1=liemp
            gquip2=liemp
            lquip1=liemp
            lquip2=liemp
            vquip1=liemp
            vquip2=liemp


            #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX 5 FORMA 5 VISITANTE
        try:
            print(dagen[0])
            dagec1  = dagen[0].split('|')
            dagenc2 = dagec1[1].split('-')
            print(dagenc2)
            dagenc3 = dagenc2[0].replace(' ','')
            print(dagenc3)
            print(cquip1[6].replace(' ',''))
            c=cquip1[6].replace(' ','')
            c1=c.split('-')
            g=gquip1[6].replace(' ','')
            g1=g.split('-')
            l=lquip1[6].replace(' ','')
            l1=l.split('-')
            v=vquip1[6].replace(' ','')
            v1=v.split('-')

            print(dagenc3)
            print(c1[0])

            if dagenc3 == c1[0]:
                caden1 = cquip1+cquip2
            else:
                caden1 = cquip2+cquip1

            if dagenc3 == g1[0]:
                caden2 = gquip1+gquip2
            else:
                caden2 = gquip2+gquip1

            if dagenc3 == l1[0]:
                caden3 = lquip1+lquip2
            else:
                caden3 = lquip2+lquip1

            if dagenc3 == v1[0]:
                caden4 = vquip1+vquip2
            else:
                caden4 = vquip2+vquip1

            datfl = dagen+caden1+caden2+caden3+caden4
            print('ingreso al try')
        except:
            dagen = ['NUL 0-0 NUL | Null - Null | Resumen']
            print(dagen[0])
            dagec1  = dagen[0].split('|')
            dagenc2 = dagec1[1].split('-')
            print(dagenc2)
            dagenc3 = dagenc2[0].replace(' ','')
            print(dagenc3)
            print(cquip1[6].replace(' ',''))
            c=cquip1[6].replace(' ','')
            c1=c.split('-')
            g=gquip1[6].replace(' ','')
            g1=g.split('-')
            l=lquip1[6].replace(' ','')
            l1=l.split('-')
            v=vquip1[6].replace(' ','')
            v1=v.split('-')

            print(dagenc3)
            print(c1[0])

            if dagenc3 == c1[0]:
                caden1 = cquip1+cquip2
            else:
                caden1 = cquip2+cquip1

            if dagenc3 == g1[0]:
                caden2 = gquip1+gquip2
            else:
                caden2 = gquip2+gquip1

            if dagenc3 == l1[0]:
                caden3 = lquip1+lquip2
            else:
                caden3 = lquip2+lquip1

            if dagenc3 == v1[0]:
                caden4 = vquip1+vquip2
            else:
                caden4 = vquip2+vquip1

            datfl = dagen+caden1+caden2+caden3+caden4
            print('ingreso al except')

        print('producto_final'+' tiempo_estimado')
        finish_time = time.time()
        print(datfl)
        print(start_time-finish_time)
        partido = datfl[0]+'|'+divis+'|'+str(tip)
        pj_acum_1 = datfl[1]
        g_acum_1 = datfl[2]
        e_acum_1 = datfl[3]
        p_acum_1 = datfl[4]
        dif_acum_1 = datfl[5]
        pts_acum_1 = datfl[6]
        equipo_acum_1 = datfl[7]
        pj_acum_2 = datfl[8]
        g_acum_2 = datfl[9]
        e_acum_2 = datfl[10]
        p_acum_2 = datfl[11]
        dif_acum_2 = datfl[12]
        pts_acum_2 = datfl[13]
        equipo_acum_2 = datfl[14]
        pj_general_5_1 = datfl[15]
        g_general_5_1 = datfl[16]
        e_general_5_1 = datfl[17]
        p_general_5_1 = datfl[18]
        dif_general_5_1 = datfl[19]
        pts_general_5_1 = datfl[20]
        equipo_general_5_1 = datfl[21]
        pj_general_5_2 = datfl[22]
        g_general_5_2 = datfl[23]
        e_general_5_2 = datfl[24]
        p_general_5_2 = datfl[25]
        dif_general_5_2 = datfl[26]
        pts_general_5_2 = datfl[27]
        equipo_general_5_2 = datfl[28]
        pj_local_5_1 = datfl[29]
        g_local_5_1 = datfl[30]
        e_local_5_1 = datfl[31]
        p_local_5_1 = datfl[32]
        dif_local_5_1 = datfl[33]
        pts_local_5_1 = datfl[34]
        equipo_local_5_1 = datfl[35]
        pj_local_5_2 = datfl[36]
        g_local_5_2 = datfl[37]
        e_local_5_2 = datfl[38]
        p_local_5_2 =datfl[39]
        dif_local_5_2 = datfl[40]
        pts_local_5_2 = datfl[41]
        equipo_local_5_2 = datfl[42]
        pj_visita_5_1 = datfl[43]
        g_visita_5_1 = datfl[44]
        e_visita_5_1 = datfl[45]
        p_visita_5_1 = datfl[46]
        dif_visita_5_1 = datfl[47]
        pts_visita_5_1 = datfl[48]
        equipo_visita_5_1 = datfl[49]
        pj_visita_5_2 = datfl[50]
        g_visita_5_2 = datfl[51]
        e_visita_5_2 = datfl[52]
        p_visita_5_2 = datfl[53]
        dif_visita_5_2 = datfl[54]
        pts_visita_5_2 = datfl[55]
        equipo_visita_5_2 = datfl[56]


        df = df.append({
                        'PARTIDO' :partido,
                        'PJ_ACUM_1' :pj_acum_1,
                        'G_ACUM_1' :g_acum_1,
                        'E_ACUM_1' :e_acum_1,
                        'P_ACUM_1' :p_acum_1,
                        'DIF_ACUM_1' :dif_acum_1,
                        'PTS_ACUM_1' :pts_acum_1,
                        'EQUIPO_ACUM_1' :equipo_acum_1,
                        'PJ_ACUM_2' :pj_acum_2,
                        'G_ACUM_2' :g_acum_2,
                        'E_ACUM_2' :e_acum_2,
                        'P_ACUM_2' :p_acum_2,
                        'DIF_ACUM_2' :dif_acum_2,
                        'PTS_ACUM_2' :pts_acum_2,
                        'EQUIPO_ACUM_2' :equipo_acum_2,
                        'PJ_GENERAL_5_1' :pj_general_5_1,
                        'G_GENERAL_5_1' :g_general_5_1,
                        'E_GENERAL_5_1' :e_general_5_1,
                        'P_GENERAL_5_1' :p_general_5_1,
                        'DIF_GENERAL_5_1' :dif_general_5_1,
                        'PTS_GENERAL_5_1' :pts_general_5_1,
                        'EQUIPO_GENERAL_5_1' :equipo_general_5_1,
                        'PJ_GENERAL_5_2' :pj_general_5_2,
                        'G_GENERAL_5_2' :g_general_5_2,
                        'E_GENERAL_5_2' :e_general_5_2,
                        'P_GENERAL_5_2' :p_general_5_2,
                        'DIF_GENERAL_5_2' :dif_general_5_2,
                        'PTS_GENERAL_5_2' :pts_general_5_2,
                        'EQUIPO_GENERAL_5_2' :equipo_general_5_2,
                        'PJ_LOCAL_5_1' :pj_local_5_1,
                        'G_LOCAL_5_1' :g_local_5_1,
                        'E_LOCAL_5_1' :e_local_5_1,
                        'P_LOCAL_5_1' :p_local_5_1,
                        'DIF_LOCAL_5_1' :dif_local_5_1,
                        'PTS_LOCAL_5_1' :pts_local_5_1,
                        'EQUIPO_LOCAL_5_1' :equipo_local_5_1,
                        'PJ_LOCAL_5_2' :pj_local_5_2,
                        'G_LOCAL_5_2' :g_local_5_2,
                        'E_LOCAL_5_2' :e_local_5_2,
                        'P_LOCAL_5_2' :p_local_5_2,
                        'DIF_LOCAL_5_2' :dif_local_5_2,
                        'PTS_LOCAL_5_2' :pts_local_5_2,
                        'EQUIPO_LOCAL_5_2' :equipo_local_5_2,
                        'PJ_VISITA_5_1' :pj_visita_5_1,
                        'G_VISITA_5_1' :g_visita_5_1,
                        'E_VISITA_5_1' :e_visita_5_1,
                        'P_VISITA_5_1' :p_visita_5_1,
                        'DIF_VISITA_5_1' :dif_visita_5_1,
                        'PTS_VISITA_5_1' :pts_visita_5_1,
                        'EQUIPO_VISITA_5_1' :equipo_visita_5_1,
                        'PJ_VISITA_5_2' :pj_visita_5_2,
                        'G_VISITA_5_2' :g_visita_5_2,
                        'E_VISITA_5_2' :e_visita_5_2,
                        'P_VISITA_5_2' :p_visita_5_2,
                        'DIF_VISITA_5_2' :dif_visita_5_2,
                        'PTS_VISITA_5_2' :pts_visita_5_2,
                        'EQUIPO_VISITA_5_2' :equipo_visita_5_2

                        } , ignore_index=True)

        print(conteo)
        print('Condicion de guardado')
        print( len(idpric) )

        if conteo % 100 == 0 or conteo == len(idpric):
            print('ingreso al if')
            df.to_excel( patha + file_name, sheet_name='DATA')
            print(patha + file_name)
            print('Succes in the dowload of data')
            time.sleep(segc)


            #driverc = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)


    print('Succes final in the dowload of data')
    driverc.quit()

#if 1==1:
    # Load the workbook
    wb = openpyxl.load_workbook(patha + file_name)
    # Select the active sheet
    sheet = wb['DATA']
    # Call the function to sum columns A and B and write the result in column C
    div_columns(sheet, 8, 3, 59)
    div_columns(sheet, 15, 10, 60)
    div_columns(sheet, 22, 17, 68)
    div_columns(sheet, 29, 24, 69)
    div_columns(sheet, 36, 31, 77)
    div_columns(sheet, 57, 52, 78)
    div_gc_columns(sheet, 7, 3, 63)
    div_gc_columns(sheet, 14, 10, 65)
    div_gc_columns(sheet, 21, 17, 72)
    div_gc_columns(sheet, 28, 24, 74)
    div_gc_columns(sheet, 35, 31, 81)
    div_gc_columns(sheet, 56, 52, 83)
    div_gf_columns(sheet, 7, 3, 62)
    div_gf_columns(sheet, 14, 10, 64)
    div_gf_columns(sheet, 21, 17, 71)
    div_gf_columns(sheet, 28, 24, 73)
    div_gf_columns(sheet, 28, 24, 73)
    div_gf_columns(sheet, 35, 31, 80)
    div_gf_columns(sheet, 56, 52, 82)
    div_var_columns_50(sheet, 59, 60, 59 , 61)
    div_var_columns_50(sheet, 62, 64, 62 , 66)
    div_var_columns_50(sheet, 65, 63, 65 , 67)
    div_var_columns(sheet, 68, 69, 68 , 70)
    div_var_columns(sheet, 71, 73, 71 , 75)
    div_var_columns(sheet, 74, 72, 74 , 76)
    div_var_columns(sheet, 77, 78, 77 , 79)
    div_var_columns(sheet, 80, 82, 80 , 84)
    div_var_columns(sheet, 83, 81, 83 , 85)
    same(sheet, 61, 86)
    same(sheet, 66, 87)
    same(sheet, 67, 88)
    same(sheet, 70, 89)
    same(sheet, 75, 90)
    same(sheet, 76, 91)
    same(sheet, 79, 92)
    same(sheet, 84, 93)
    same(sheet, 85, 94)
    average_dist(sheet, 61, 66, 67 , 95)
    average_dist(sheet, 70, 75, 76 , 96)
    average_dist(sheet, 79, 84, 85 , 97)
    sum_tot(sheet, 61, 66, 67, 70, 75, 76, 79, 84, 85 , 98)
    average_dist(sheet, 95, 96, 97 , 99)
    desvest(sheet, 61, 66, 67, 70, 75, 76, 79, 84, 85 , 100)
    sum_tot_abs(sheet, 61, 66, 67, 70, 75, 76, 79, 84, 85 , 101)
    average_dist_abs(sheet, 95, 96, 97 , 102)
    same(sheet, 2, 103)
    concatkeynovo(sheet, 2, '|', 104)
    valida(sheet, 104, 3, 58 , 105)
    resultado(sheet, 98, 99, 100,109)
    sheet['BG1'].value = 'PROM_ACUM_L'
    sheet['BH1'].value = 'PROM_ACUM_V'
    sheet['BI1'].value = '1_EVA'
    sheet['BJ1'].value = 'GF_LOCAL'
    sheet['BK1'].value = 'GC_LOCAL'
    sheet['BL1'].value = 'GF_VISITA'
    sheet['BM1'].value = 'GC_VISITA'
    sheet['BN1'].value = '2_EVA'
    sheet['BO1'].value = '3_EVA'
    sheet['BP1'].value = 'PROM_ACUM_L_F5'
    sheet['BQ1'].value = 'PROM_ACUM_V_F5'
    sheet['BR1'].value = '1_EVA_F5'
    sheet['BS1'].value = 'GF_LOCAL_F5'
    sheet['BT1'].value = 'GC_LOCAL_F5'
    sheet['BU1'].value = 'GF_VISITA_F5'
    sheet['BV1'].value = 'GC_VISITA_F5'
    sheet['BW1'].value = '2_EVA_F5'
    sheet['BX1'].value = '3_EVA_F5'
    sheet['BY1'].value = 'PROM_ACUM_L_F5'
    sheet['BZ1'].value = 'PROM_ACUM_V_F5'
    sheet['CA1'].value = '1_EVA_F5'
    sheet['CB1'].value = 'GF_LOCAL_F5'
    sheet['CC1'].value = 'GC_LOCAL_F5'
    sheet['CD1'].value = 'GF_VISITA_F5'
    sheet['CE1'].value = 'GC_VISITA_F5'
    sheet['CF1'].value = '2_EVA_F5'
    sheet['CG1'].value = '3_EVA_F5'
    sheet['CH1'].value = 'EVA1'
    sheet['CI1'].value = 'EVA2'
    sheet['CJ1'].value = 'EVA3'
    sheet['CK1'].value = 'EVA4'
    sheet['CL1'].value = 'EVA5'
    sheet['CM1'].value = 'EVA6'
    sheet['CN1'].value = 'EVA7'
    sheet['CO1'].value = 'EVA8'
    sheet['CP1'].value = 'EVA9'
    sheet['CQ1'].value = 'AVER1'
    sheet['CR1'].value = 'AVER2'
    sheet['CS1'].value = 'AVER3'
    sheet['CT1'].value = 'TOTAL'
    sheet['CU1'].value = 'AVER3'
    sheet['CV1'].value = 'DEV1'
    sheet['CW1'].value = 'TOV2'
    sheet['CX1'].value = 'AVV3'
    sheet['CY1'].value = 'RES'
    sheet['CZ1'].value = 'VALIDA_LOCAL'
    sheet['DA1'].value = 'ANALIZA'
    sheet['DB1'].value = 'RESULTADO'
    sheet['DC1'].value = 'CUENTA'
    sheet['DD1'].value = 'RESR'
    sheet['DE1'].value = 'RESF'
    sheet['DF1'].value = 'COINCIDIO'
    sheet['DG1'].value = 'CHOICE'
    sheet['DH1'].value = 'EFIC_1'
    sheet['DI1'].value = 'EFIC_2'
    sheet['DJ1'].value = 'KORT3'
    sheet['DK1'].value = 'KORT4'
    sheet['DL1'].value = 'KEY_LOG'
    sheet['DM1'].value = 'KORT5'
    sheet['DN1'].value = 'KEY_CUO'

    sheet['A1'].value  = 'INDEX'

    assetf = filter_column(sheet, 100, 103, 102, 109, 105, 104, 0.5, 0.5, 'SI')
    assetx = filter_column(sheet, 100, 103, 102, 109, 105, 104, 5, 0.1, 'SI')
    #ssetf = filter_column(sheet, 100, 103, 102, 109, 105, 104, 3, 0.1, 'SI')
    # Save the workbook
    wb.save(patha + file_name)
    wb.close()
    print(assetf)
    #table = [assetf[i:i+3] for i in range(0, len(assetf), 3)]
    #table_string = tabulate(table, headers=['Datos_Generales', 'Resultado_Esperado', 'Local_Visita'], tablefmt='grid')
    sub_lists = [assetf[i:i+3] for i in range(0, len(assetf), 3)]
    result = '[' + '], ['.join([str(sub_list) for sub_list in sub_lists]) + ']'
    string_result = str(result.encode('utf-8'))
    sub_listx = [assetx[i:i+3] for i in range(0, len(assetx), 3)]
    resultx = '[' + '], ['.join([str(sub_listx) for sub_list in sub_listx]) + ']'
    string_resultx = str(resultx.encode('utf-8'))

    time.sleep(segc)
    pyautogui.hotkey('win','r')
    pyautogui.write(pathd + 'betdash.xlsx' ,interval = 0.5)
    pyautogui.press('enter',interval = 0.5)
    time.sleep(segc)
    time.sleep(segc)
    pyautogui.hotkey('ctrl','alt','f5',interval = 0.5)
    time.sleep(120)
    pyautogui.hotkey('ctrl','s',interval = 0.5)
    time.sleep(segc)
    pyautogui.hotkey('alt','f4',interval = 0.5)

#if 1==1:
    print(patha + file_name)
    print(pathd + file_named)
    wbx = openpyxl.load_workbook(patha + file_name)

    try:
        sheet_to_remove = wbx['HIST_EFIC']
        sheet_to_removc = wbx['HIST_CUOT']
        wbx.remove(sheet_to_remove)
        wbx.remove(sheet_to_removc)
        wbx_efic = wbx.create_sheet(title = 'HIST_EFIC') #para copiar la eficiencia
        wbx_cuot = wbx.create_sheet(title = 'HIST_CUOT') #para copiar la ccuota
    except:
        wbx_efic = wbx.create_sheet(title = 'HIST_EFIC') #para copiar la eficiencia
        wbx_cuot = wbx.create_sheet(title = 'HIST_CUOT') #para copiar la ccuota
    data_final = wbx['DATA']
    dbx = openpyxl.load_workbook(pathd + file_named)
    dbx_efic = dbx['EFICIENCIA']
    dbx_cuot = dbx['CUOTA']

    copy_column(dbx_efic, wbx_efic, 1, 1)
    copy_column(dbx_efic, wbx_efic, 2, 2)
    copy_column(dbx_efic, wbx_efic, 3, 3)
    copy_column(dbx_efic, wbx_efic, 4, 4)
    copy_column(dbx_efic, wbx_efic, 5, 5)
    copy_column(dbx_efic, wbx_efic, 6, 6)
    same(wbx_efic, 5, 8)
    wbx_efic['G1'].value = 'LLAVE'
    wbx_efic['H1'].value = '%'
    concatcuo(wbx_efic, 1, 2,  7)

    copy_column(dbx_cuot, wbx_cuot, 1, 1)
    copy_column(dbx_cuot, wbx_cuot, 2, 2)
    copy_column(dbx_cuot, wbx_cuot, 3, 3)
    copy_column(dbx_cuot, wbx_cuot, 4, 4)
    copy_column(dbx_cuot, wbx_cuot, 5, 5)
    copy_column(dbx_cuot, wbx_cuot, 6, 6)
    same(wbx_cuot, 5, 8)
    wbx_cuot['G1'].value = 'LLAVE'
    wbx_cuot['H1'].value = '%'
    concatcuo(wbx_cuot, 1, 2,  7)

    resul_cuo(data_final, 112)
    resul_log(data_final, 113)
    kort3(data_final,114)
    kort4(data_final,115)
    keylog(data_final,116)
    kort5(data_final,117)
    keycuo(data_final,118)
    wbx.save(patha + file_name)
    #dbx.save(pathd + file_named)
    wbx.close()
    dbx.close()


#if 1==1:
    # Load the original workbook
    workbook = openpyxl.load_workbook(patha + file_name)
    # Create a copy of the workbook
    copy_workbook = openpyxl.Workbook()
    # Copy the contents of the original workbook to the copy workbook
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        copy_worksheet = copy_workbook.create_sheet(sheet_name)
        for row in worksheet.iter_rows():
            copy_row = [cell.value for cell in row]
            copy_worksheet.append(copy_row)
    # Get the sheet you want to modify
    sheet = copy_workbook['DATA']
    #sheet.delete_cols(2, 103, shiftCols=False)
    # Loop through columns 2 to 4 and delete the contents of each cell
    for column in range(2, 102):
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=column)
            cell.value = None

    sheet_to_remove = copy_workbook['Sheet']
    copy_workbook.remove(sheet_to_remove)
    # Save the modified workbook

    copy_workbook.save(pathc + file_namec)
    workbook.close()
    copy_workbook.close()

    wx = openpyxl.load_workbook(pathc + file_namec)
    wy = wx.active
    #wy.delete_cols(2, 98)
    wx.save(pathc + file_namec)
    #wy.delete_cols(14, 21)
    wy.column_dimensions.group(start='B', end='CX', hidden=True)
    wx.save(pathc + file_namec)
    wx.close()
    #wy.save(pathc + file_namec)
    #wy.close()
    #carta=''
    #for letter in assetf:
    #    carta = carta   + letter + '|'
    #carta_utf = table_string.encode('utf-8')
    #print(carta)
    #send_email_x(subject, string_result , to_email, from_email, password)


    send_email_y(subject, message, to_email, from_email, password,patha + file_name,file_name)
    print('Se envio el correo admin')
    time.sleep(segc)
    send_email_y(subject, message, to_email, from_email, password,pathc + file_namec,file_namec)


    send_email_y(subject, message, 'rtriguerosvargas@gmail.com', from_email, password,pathc + file_namec,file_namec)
    send_email_y(subject, message, 'Gustavocarazasc@gmail.com', from_email, password,pathc + file_namec,file_namec)
    send_email_y(subject, message, 'Willmezamotta@gmail.com', from_email, password,pathc + file_namec,file_namec)
    send_email_y(subject, message, 'Ricardo.950219@gmail.com', from_email, password,pathc + file_namec,file_namec)
    send_email_y(subject, message, 'a.escrib137@gmail.com', from_email, password,pathc + file_namec,file_namec)
    send_email_y(subject, message, 'Zadihr90@gmail.com', from_email, password,pathc + file_namec,file_namec)
    send_email_y(subject, message, 'mamoblitas@gmail.com', from_email, password,pathc + file_namec,file_namec)
    send_email_y(subject, message, 'jhmciurliza@gmail.com', from_email, password,pathc + file_namec,file_namec)
    print('Se envio el correo a clientes')
    time.sleep(segc)



#if 1==1:

    #pyautogui.hotkey('win','r')
    #pyautogui.write(patha + file_name ,interval = 0.5)
    #pyautogui.press('enter',interval = 0.5)
    #time.sleep(segc)
    #pyautogui.hotkey('ctrl','s',interval = 0.5)
    #time.sleep(segc)
    #pyautogui.hotkey('alt','f4',interval = 0.5)
    #time.sleep(segc)
    #time.sleep(segc)
    #pyautogui.hotkey('win','r')
    #pyautogui.write(pathd + 'betdash.xlsx' ,interval = 0.5)
    #pyautogui.press('enter',interval = 0.5)
    #time.sleep(segc)
    #time.sleep(segc)
    #pyautogui.hotkey('ctrl','alt','f5',interval = 0.5)
    #time.sleep(80)
    #pyautogui.hotkey('alt','f4',interval = 0.5)



except Exception as e:

    message = f"Error occurred: {str(e)}"
    trace = traceback.format_exc()
    x = f"{message}\n{trace}"
    send_email_x('Error', str(x)  , to_email, from_email, password)
    #logger.exception("Exception Occured while code Execution: "+ str(e))
    print(x)
    print('Se envio el correo de error')


#res= 'AUG - BRE | Augsburgo - Werder Bremen | Resumen|ALEMANIA: Bundesliga - Jornada 23|2.40'


#print(res[12:37]+'s')
