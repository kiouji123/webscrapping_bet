# -*- coding: utf-8 -*-
"""
Created on Sat Mar 14 00:31:36 2020

@author: Juan Renatto
"""
from selenium import webdriver
import pandas as pd
from bs4 import BeautifulSoup
import requests
import time
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date, timedelta
import smtplib
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import statistics
import logging
import traceback
from tabulate import tabulate
import pyautogui
#SELENIUM ---------------------------XXXXXXXXXXXXXXXXX
pyautogui.FAILSAFE = False
urlply= "https://www.flashscore.pe/futbol/peru/liga-1/resultados/"
executable_path='C:\Program Files (x86)\chromedriver.exe'
liemp = ['0','0','0','0','0:0','0','NULL']
stodat = date.today()
ayer = False
if ayer== True:
    today = stodat  - timedelta(days=1)
    print('busca ayer')
else:
    today = stodat
    print('busca hoy')

patha = 'C:/Users/mmo_2/OneDrive/oc/resultados/'
pathb = 'C:/Users/mmo_2/OneDrive/oc/data/'
file_name = 'res_' + str(today) + '.xlsx'
file_nameb = 'data_' + str(today) + '.xlsx'
file_namec = 'resclie_' + str(today) + '.xlsx'
print('nombre del archivo')
print(file_name)
pyautogui.PAUSE = 2.5

#def send_email(subject, message, to_email, from_email, password):
#    try:
#        server = smtplib.SMTP('smtp.gmail.com', 587)
#        server.ehlo()
#        server.starttls()
#        server.login(from_email, password)
#        message = f"Subject: {subject}\n\n{message}"
#        server.sendmail(from_email, to_email, message)
#        print("Email sent!")
#    except Exception as e:
#        print(f"Something went wrong... {e}")
#    finally:
#        server.quit()

def send_email_x(subject, message, to_email, from_email, password):
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(from_email, password)
        message = f"Subject: {subject}\n\n{message}"
        server.sendmail(from_email, to_email, message)
        print("Email sent!")
    except Exception as e:
        print(f"Something went wrong... {e}")
    finally:
        server.quit()

def send_email_y(subject, message, to_email, from_email, password, file_path, workbname):
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(from_email, password)

        # Create a multipart message
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject

        # Add the message body
        body = MIMEText(message)
        msg.attach(body)

        # Add the Excel file as an attachment
        with open(file_path, 'rb') as f:
            attachment = MIMEApplication(f.read(), _subtype='xlsx')
            attachment.add_header('Content-Disposition', 'attachment', filename=workbname)
            msg.attach(attachment)

        # Send the email
        server.sendmail(from_email, to_email, msg.as_string())
        print("Email sent!")
    except Exception as e:
        print(f"Something went wrong... {e}")
    finally:
        server.quit()

subject = "Resutado del dia"
message = "Porcentaje y resultados del dia generales y especificos."
to_email = "juanre_68@hotmail.com"
from_email = "juanre060894@gmail.com"
password = "jdlwjzmbahaxvbwq"




def has_id(tag):
    return tag.has_attr('id')


def has_name(tag):
    return tag.has_attr('name')


options = Options()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_argument('--ignore-certificate-errors-spki-list')
options.add_argument('--headless')

segs = random.randrange(5, 8)
segc = random.randrange(4, 6)
sego = random.randrange(4, 6)

def copy_column(src_sheet, dst_sheet, src_col, dst_col):
    for row in src_sheet.iter_rows(values_only=True):
        dst_sheet.cell(row=dst_sheet.max_row + 1, column=dst_col).value = row[src_col - 1]

def find_substring_1(string, substring, start):
    return int(string.find(substring, start - 1) + 1)

def find_substring_2(string, substring, start):
    return int(string.find(substring, start)+1)


def choice_matchs(sheet, filt1, filt2, filt3,col_result, threshold_1 , threshold_2, threshold_3):

    for row in range(2, sheet.max_row + 1):
        desve1 = sheet.cell(row=row, column=filt1)
        averg1 = sheet.cell(row=row, column=filt2)
        opinio = sheet.cell(row=row, column=filt3)
        if desve1.value is not None and averg1.value is not None and opinio.value is not None:
            if desve1.value <= threshold_1 and averg1.value >= threshold_2 and opinio.value == threshold_3:
                result = 'CHOICE'
                result_cell = sheet.cell(row=row, column=col_result)
                result_cell.value = result
            else:
                result_cell = sheet.cell(row=row, column=col_result)
                result_cell.value = 0




def realidad(f, g):
    if f > g:
        return "LOCAL"
    elif g > f:
        return "VISITA"
    else:
        return "EMPATE"

def match_result(f2, g2):
    if abs(f2 - g2) >= 3:
        return "GOLEADA"
    else:
        return "NORMAL"







def corte_1(sheet, col1, result_col):
    for row in range(3, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        try:
            result = find_substring_1(cell1.value, '|', 1)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def corte_2(sheet, col1,col2, result_col):
    for row in range(3, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        try:
            result = find_substring_2(cell1.value, '|', int(cell2.value))
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def llave_res(sheet, col1,col2,col3, result_col):
    for row in range(3, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        cell3 = sheet.cell(row=row, column=col3)
        try:
            result = cell1.value[ int(cell2.value)+1: int(cell3.value)-2]
            print('result')
            print(result)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def corte_3(sheet, col1,col2, result_col):
    for row in range(3, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        try:
            result = cell1.value[0:int(cell2.value)-2]
            print(result)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def local(sheet, col1, result_col):
    for row in range(3, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)

        try:
            print('find')
            print(find_substring_2(cell1.value, '-', 1)-1)
            result =  cell1.value[(find_substring_1(cell1.value, '-', 1)-2):find_substring_1(cell1.value, '-', 1)-1]
            print('local')
            print(result)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def visita(sheet, col1, result_col):
    for row in range(3, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        try:
            print('find')
            print(find_substring_2(cell1.value, '-', 1)-1)
            result = cell1.value[(find_substring_1(cell1.value, '-', 1)):find_substring_1(cell1.value, '-', 1)+1]
            print('visita')
            print(result)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        except:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0

def resul_final(sheet, col1,col2, result_col):
    for row in range(3, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        if cell1.value != ' ' and cell2.value !=' ':
            result = realidad(int(cell1.value),int(cell2.value))
            print('resultado del partido')
            print(result)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        else:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = 0


def goleada(sheet, col1,col2, result_col):
    for row in range(3, sheet.max_row + 1):
        cell1 = sheet.cell(row=row, column=col1)
        cell2 = sheet.cell(row=row, column=col2)
        if cell1.value != ' ' and cell2.value !=' ':
            result = match_result(int(cell1.value),int(cell2.value))
            print('resultado del partido')
            print(result)
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
        else:
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = ''


def resulx(sheet, result_col):
    conteo = 2
    for row in range(conteo, sheet.max_row + 1):
            result =  '=IF(ISNA(VLOOKUP(CZ'+str(conteo)+',ANALISIS!D:E,2,FALSE)),"",VLOOKUP(CZ'+str(conteo)+',ANALISIS!D:E,2,FALSE))'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1


def cuenta(sheet, result_col):
    conteo = 2
    for row in range(conteo, sheet.max_row + 1):
            result =  '=IF(AND(DB'+str(conteo)+'<>"",DA'+str(conteo)+'="SI"),"SI","NO")'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1

def resr(sheet, result_col):
    conteo = 2
    for row in range(2, sheet.max_row + 1):
            result =  '=IF(ISNA(VLOOKUP(CZ'+str(conteo)+',ANALISIS!D:H,5,FALSE)),"",VLOOKUP(CZ'+str(conteo)+',ANALISIS!D:H,5,FALSE))'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1

def coinci(sheet, result_col):
    conteo = 2
    for row in range(2, sheet.max_row + 1):
            result =  '=IF(OR(DD'+str(conteo)+'="",DD'+str(conteo)+'=0,DC'+str(conteo)+'="NO"),3,IF(DD'+str(conteo)+'=DE'+str(conteo)+',1,IF(AND(DD'+str(conteo)+'="VISITA",DE'+str(conteo)+'="VISITA-EMPATE"),1,IF(AND(DD'+str(conteo)+'="LOCAL",DE'+str(conteo)+'="LOCAL-EMPATE"),1,IF(AND(DD'+str(conteo)+'="EMPATE",OR(DE'+str(conteo)+'="LOCAL-EMPATE",DE'+str(conteo)+'="VISITA-EMPATE")),1,0)))))'
            result_cell = sheet.cell(row=row, column=result_col)
            result_cell.value = result
            conteo = conteo+1



df = pd.DataFrame(
    columns=
            [
                'PARTIDO',
                'PJ_ACUM_1',
                'G_ACUM_1',
                'E_ACUM_1',
                'P_ACUM_1',
                'DIF_ACUM_1',
                'PTS_ACUM_1',
                'EQUIPO_ACUM_1',
                'PJ_ACUM_2',
                'G_ACUM_2',
                'E_ACUM_2',
                'P_ACUM_2',
                'DIF_ACUM_2',
                'PTS_ACUM_2',
                'EQUIPO_ACUM_2',
                'PJ_GENERAL_5_1',
                'G_GENERAL_5_1',
                'E_GENERAL_5_1',
                'P_GENERAL_5_1',
                'DIF_GENERAL_5_1',
                'PTS_GENERAL_5_1',
                'EQUIPO_GENERAL_5_1',
                'PJ_GENERAL_5_2',
                'G_GENERAL_5_2',
                'E_GENERAL_5_2',
                'P_GENERAL_5_2',
                'DIF_GENERAL_5_2',
                'PTS_GENERAL_5_2',
                'EQUIPO_GENERAL_5_2',
                'PJ_LOCAL_5_1',
                'G_LOCAL_5_1',
                'E_LOCAL_5_1',
                'P_LOCAL_5_1',
                'DIF_LOCAL_5_1',
                'PTS_LOCAL_5_1',
                'EQUIPO_LOCAL_5_1',
                'PJ_LOCAL_5_2',
                'G_LOCAL_5_2',
                'E_LOCAL_5_2',
                'P_LOCAL_5_2',
                'DIF_LOCAL_5_2',
                'PTS_LOCAL_5_2',
                'EQUIPO_LOCAL_5_2',
                'PJ_VISITA_5_1',
                'G_VISITA_5_1',
                'E_VISITA_5_1',
                'P_VISITA_5_1',
                'DIF_VISITA_5_1',
                'PTS_VISITA_5_1',
                'EQUIPO_VISITA_5_1',
                'PJ_VISITA_5_2',
                'G_VISITA_5_2',
                'E_VISITA_5_2',
                'P_VISITA_5_2',
                'DIF_VISITA_5_2',
                'PTS_VISITA_5_2',
                'EQUIPO_VISITA_5_2',




             ]
              )

try:

    print(today)
    print(stodat)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
    driver.get("https://www.flashscore.pe/?rd=mismarcadores.com")

    time.sleep(segs)
    if ayer == True:
        buttoy = driver.find_element(By.XPATH,"//button[@class='calendar__navigation calendar__navigation--yesterday']" )
        driver.execute_script("arguments[0].click();", buttoy)
        print('fue al dia de ayer')
    #button = driver.find_element(By.XPATH, '/html/body/div[4]/div[1]/div/div/main/div[4]/div[2]/div/section/div/div/div[1]/span')
    #buttons = driver.find_elements(By.XPATH, "//span[@title='Ocultar todos los partidos de esta liga']")
    #print('button')
    #print(button)
    #print('buttons')
    #print(buttons)
    #button.click()


    number_li_elems=len(WebDriverWait(driver,30).until(EC.visibility_of_all_elements_located((By.XPATH, "//span[@title='Mostrar todos los partidos de esta liga']"))))
    print(number_li_elems)
    #for x in range(number_li_elems):
    for x in range(number_li_elems):
        time.sleep(sego)
        # you have to get the element by index every time, otherwise you will get StaleElement Exception
        #button = driver.find_element(By.XPATH, "//span[@title='Ocultar todos los partidos de esta liga']")
        button = driver.find_element(By.XPATH, "//span[@title='Mostrar todos los partidos de esta liga']")
        #Mostrar todos los partidos de esta liga
        print(button)
        driver.execute_script("arguments[0].click();", button)




    time.sleep(segs)
    html = driver.page_source
    soupply= BeautifulSoup (html,'html.parser')
    tableply = soupply.find('div',class_ ='sportName soccer') #, id='stats_standard'
    #print(tableply)
    tableplx = tableply.find_all(has_id)
    #tableplx = tableply.find_all('div',class_ ='event__match event__match--twoLine')

    print('CORTE')
    #print(tableplx)

    #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX

    idprin = []




    for idx in tableplx:
        idclea = idx.get('id')
        idprin.append(idclea[4:len(idclea)])
        #print(idx.get('id'))
    # http://example.com/elsie
    # http://example.com/lacie
    # http://example.com/tillie
    print(idprin)

    #--------------------------VAMOS
    time.sleep(segc)

    buttob = driver.find_element(By.XPATH,"//div[text()='Cuotas']")

    driver.execute_script("arguments[0].click();", buttob)

    time.sleep(segc)
    time.sleep(segc)
    time.sleep(segc)
    htmlb = driver.page_source

    soupplyb= BeautifulSoup (htmlb,'html.parser')
    idpric = []
    driver.quit()

    for idc in idprin:
        idcx= 'g_1_' + idc
        tableplb = soupplyb.find('div',id =idcx) #, id='stats_standard'
        print(tableplb)
        #tablepbx = tableplb.find('div', class_=re.compile('^odds__odd event__odd--odd1'))
        tablepbx=tableplb.find("div", class_=lambda value: value and value.startswith("odds__odd event__odd--odd1"))

        print(tablepbx)
        if tablepbx.text != '-':
            idpric.append(idc)
            print('con cuota')
        else:
            print('sin cuota')


    print('CORTE CUOTA')
    print(idpric)
    print('CORTE CUOTA')


    #--------------------------VAMOS



    #for urlx in resumx[:10]:

    #    segs = random.randrange(10, 13)
    #    time.sleep(segs)
    #    print('ingreso')
    #    source=requests.get(urlx,timeout=5 ).text
    #    soup= BeautifulSoup (source,'html.parser')
    #    table = soup.title.string#

    conteo = 0
    totalp = len(idpric)
    driverc = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
    for key in idpric:

        resume = 'https://www.flashscore.pe/partido/'+key+'/#/resumen-del-partido/resumen-del-partido'
        clasif = 'https://www.flashscore.pe/partido/'+key+'/#/clasificacion/table/overall'
        formag = 'https://www.flashscore.pe/partido/'+key+'/#/clasificacion/form/overall/5'
        formal = 'https://www.flashscore.pe/partido/'+key+'/#/clasificacion/form/home/5'
        formav = 'https://www.flashscore.pe/partido/'+key+'/#/clasificacion/form/away/5'
        #resumx.append(resume)
        #clasix.append(clasif)
        #formgx.append(formag)
        #formlx.append(formal)
        #formvx.append(formav)
        conteo = conteo + 1
        print(str(conteo)+'/'+str(totalp))
        #BS4 ---------------------------XXXXXXXXXXXXXXXXX 1 RECUPERACION DE EQUIPOS
        dagen = []
        time.sleep(segs)
        print('ingreso')


        driverc.get(resume)
        time.sleep(segs)
        sourcr = driverc.page_source
        #sourcr=requests.get(resume,timeout=10 ).text
        soupr= BeautifulSoup (sourcr,'html.parser')

        try:
            table = soupr.head.title.string
        except:
            table = 'NUL 0-0 NUL | Null - Null | Resumen'

        #sourcr=requests.get(resume,timeout=10 ).text
        #soupr= BeautifulSoup (sourcr,'html.parser')
        #print(soupr)
        #table = soupr.head.title.string

        meta_tag = soupr.find('meta', attrs={'name': 'og:description'})
        divis = meta_tag['content']
        #liga = divis["content"]

        print('1ER NIVEL')
        print(table)
        print(divis)
        print('1ER NIVEL')
        dagen.append(table)

        #BS4 ---------------------------XXXXXXXXXXXXXXXXX 1 RECUPERACION DE EQUIPOS

        #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX 2 RECUPERACION DE PUNTOS TOTALES LOCAL Y VISITA
        #driverc.quit()
        #time.sleep(segs)
        cquip1=liemp
        cquip2=liemp
        gquip1=liemp
        gquip2=liemp
        lquip1=liemp
        lquip2=liemp
        vquip1=liemp
        vquip2=liemp
        try:
            print(dagen[0])
            dagec1 = dagen[0].split('|')
            dagenc2 = dagec1[1].split('-')
            print(dagenc2)
            dagenc3 = dagenc2[0].replace(' ','')
            print(dagenc3)
            print(cquip1[6].replace(' ',''))
            c=cquip1[6].replace(' ','')
            c1=c.split('-')
            g=gquip1[6].replace(' ','')
            g1=g.split('-')
            l=lquip1[6].replace(' ','')
            l1=l.split('-')
            v=vquip1[6].replace(' ','')
            v1=v.split('-')

            print(dagenc3)
            print(c1[0])

            if dagenc3 == c1[0]:
                caden1 = cquip1+cquip2
            else:
                caden1 = cquip2+cquip1

            if dagenc3 == g1[0]:
                caden2 = gquip1+gquip2
            else:
                caden2 = gquip2+gquip1

            if dagenc3 == l1[0]:
                caden3 = lquip1+lquip2
            else:
                caden3 = lquip2+lquip1

            if dagenc3 == v1[0]:
                caden4 = vquip1+vquip2
            else:
                caden4 = vquip2+vquip1

            datfl = dagen+caden1+caden2+caden3+caden4
            print('ingreso al try')
        except:
            dagen = ['NUL 0-0 NUL | Null - Null | Resumen']
            print(dagen[0])
            dagec1 = dagen[0].split('|')
            dagenc2 = dagec1[1].split('-')
            print(dagenc2)
            dagenc3 = dagenc2[0].replace(' ','')
            print(dagenc3)
            print(cquip1[6].replace(' ',''))
            c=cquip1[6].replace(' ','')
            c1=c.split('-')
            g=gquip1[6].replace(' ','')
            g1=g.split('-')
            l=lquip1[6].replace(' ','')
            l1=l.split('-')
            v=vquip1[6].replace(' ','')
            v1=v.split('-')

            print(dagenc3)
            print(c1[0])

            if dagenc3 == c1[0]:
                caden1 = cquip1+cquip2
            else:
                caden1 = cquip2+cquip1

            if dagenc3 == g1[0]:
                caden2 = gquip1+gquip2
            else:
                caden2 = gquip2+gquip1

            if dagenc3 == l1[0]:
                caden3 = lquip1+lquip2
            else:
                caden3 = lquip2+lquip1

            if dagenc3 == v1[0]:
                caden4 = vquip1+vquip2
            else:
                caden4 = vquip2+vquip1

            datfl = dagen+caden1+caden2+caden3+caden4
            print('ingreso al except')
        print('producto_final')
        print(datfl)

        partido = datfl[0]+'|'+divis
        pj_acum_1 = datfl[1]
        g_acum_1 = datfl[2]
        e_acum_1 = datfl[3]
        p_acum_1 = datfl[4]
        dif_acum_1 = datfl[5]
        pts_acum_1 = datfl[6]
        equipo_acum_1 = datfl[7]
        pj_acum_2 = datfl[8]
        g_acum_2 = datfl[9]
        e_acum_2 = datfl[10]
        p_acum_2 = datfl[11]
        dif_acum_2 = datfl[12]
        pts_acum_2 = datfl[13]
        equipo_acum_2 = datfl[14]
        pj_general_5_1 = datfl[15]
        g_general_5_1 = datfl[16]
        e_general_5_1 = datfl[17]
        p_general_5_1 = datfl[18]
        dif_general_5_1 = datfl[19]
        pts_general_5_1 = datfl[20]
        equipo_general_5_1 = datfl[21]
        pj_general_5_2 = datfl[22]
        g_general_5_2 = datfl[23]
        e_general_5_2 = datfl[24]
        p_general_5_2 = datfl[25]
        dif_general_5_2 = datfl[26]
        pts_general_5_2 = datfl[27]
        equipo_general_5_2 = datfl[28]
        pj_local_5_1 = datfl[29]
        g_local_5_1 = datfl[30]
        e_local_5_1 = datfl[31]
        p_local_5_1 = datfl[32]
        dif_local_5_1 = datfl[33]
        pts_local_5_1 = datfl[34]
        equipo_local_5_1 = datfl[35]
        pj_local_5_2 = datfl[36]
        g_local_5_2 = datfl[37]
        e_local_5_2 = datfl[38]
        p_local_5_2 = datfl[39]
        dif_local_5_2 = datfl[40]
        pts_local_5_2 = datfl[41]
        equipo_local_5_2 = datfl[42]
        pj_visita_5_1 = datfl[43]
        g_visita_5_1 = datfl[44]
        e_visita_5_1 = datfl[45]
        p_visita_5_1 = datfl[46]
        dif_visita_5_1 = datfl[47]
        pts_visita_5_1 = datfl[48]
        equipo_visita_5_1 = datfl[49]
        pj_visita_5_2 = datfl[50]
        g_visita_5_2 = datfl[51]
        e_visita_5_2 = datfl[52]
        p_visita_5_2 = datfl[53]
        dif_visita_5_2 = datfl[54]
        pts_visita_5_2 = datfl[55]
        equipo_visita_5_2 = datfl[56]


        df = df.append({
                        'PARTIDO' :partido,
                        'PJ_ACUM_1' :pj_acum_1,
                        'G_ACUM_1' :g_acum_1,
                        'E_ACUM_1' :e_acum_1,
                        'P_ACUM_1' :p_acum_1,
                        'DIF_ACUM_1' :dif_acum_1,
                        'PTS_ACUM_1' :pts_acum_1,
                        'EQUIPO_ACUM_1' :equipo_acum_1,
                        'PJ_ACUM_2' :pj_acum_2,
                        'G_ACUM_2' :g_acum_2,
                        'E_ACUM_2' :e_acum_2,
                        'P_ACUM_2' :p_acum_2,
                        'DIF_ACUM_2' :dif_acum_2,
                        'PTS_ACUM_2' :pts_acum_2,
                        'EQUIPO_ACUM_2' :equipo_acum_2,
                        'PJ_GENERAL_5_1' :pj_general_5_1,
                        'G_GENERAL_5_1' :g_general_5_1,
                        'E_GENERAL_5_1' :e_general_5_1,
                        'P_GENERAL_5_1' :p_general_5_1,
                        'DIF_GENERAL_5_1' :dif_general_5_1,
                        'PTS_GENERAL_5_1' :pts_general_5_1,
                        'EQUIPO_GENERAL_5_1' :equipo_general_5_1,
                        'PJ_GENERAL_5_2' :pj_general_5_2,
                        'G_GENERAL_5_2' :g_general_5_2,
                        'E_GENERAL_5_2' :e_general_5_2,
                        'P_GENERAL_5_2' :p_general_5_2,
                        'DIF_GENERAL_5_2' :dif_general_5_2,
                        'PTS_GENERAL_5_2' :pts_general_5_2,
                        'EQUIPO_GENERAL_5_2' :equipo_general_5_2,
                        'PJ_LOCAL_5_1' :pj_local_5_1,
                        'G_LOCAL_5_1' :g_local_5_1,
                        'E_LOCAL_5_1' :e_local_5_1,
                        'P_LOCAL_5_1' :p_local_5_1,
                        'DIF_LOCAL_5_1' :dif_local_5_1,
                        'PTS_LOCAL_5_1' :pts_local_5_1,
                        'EQUIPO_LOCAL_5_1' :equipo_local_5_1,
                        'PJ_LOCAL_5_2' :pj_local_5_2,
                        'G_LOCAL_5_2' :g_local_5_2,
                        'E_LOCAL_5_2' :e_local_5_2,
                        'P_LOCAL_5_2' :p_local_5_2,
                        'DIF_LOCAL_5_2' :dif_local_5_2,
                        'PTS_LOCAL_5_2' :pts_local_5_2,
                        'EQUIPO_LOCAL_5_2' :equipo_local_5_2,
                        'PJ_VISITA_5_1' :pj_visita_5_1,
                        'G_VISITA_5_1' :g_visita_5_1,
                        'E_VISITA_5_1' :e_visita_5_1,
                        'P_VISITA_5_1' :p_visita_5_1,
                        'DIF_VISITA_5_1' :dif_visita_5_1,
                        'PTS_VISITA_5_1' :pts_visita_5_1,
                        'EQUIPO_VISITA_5_1' :equipo_visita_5_1,
                        'PJ_VISITA_5_2' :pj_visita_5_2,
                        'G_VISITA_5_2' :g_visita_5_2,
                        'E_VISITA_5_2' :e_visita_5_2,
                        'P_VISITA_5_2' :p_visita_5_2,
                        'DIF_VISITA_5_2' :dif_visita_5_2,
                        'PTS_VISITA_5_2' :pts_visita_5_2,
                        'EQUIPO_VISITA_5_2' :equipo_visita_5_2

                        } , ignore_index=True)

    if conteo % 100 == 0 or conteo == len(idpric):
        print('ingreso al if')
        df.to_excel( patha + file_name, sheet_name='RES')
        print(patha + file_name)
        print('Succes in the dowload of data')
        time.sleep(segc)
    driverc.quit()
    #df.to_excel( patha + file_name, sheet_name='RES')
    print('Succes in the dowload of data')



    # Load the source Excel file
    src_wb = openpyxl.load_workbook(patha + file_name)
    # Get the sheet you want to copy
    src_sheet = src_wb['RES']
    # Load the destination Excel file
    dst_wb = openpyxl.load_workbook(pathb + file_nameb)
    # Create a new sheet in the destination Excel file

    try:
        sheet_to_remove = dst_wb['ANALISIS']
        dst_wb.remove(sheet_to_remove)
        dst_sheet = dst_wb.create_sheet(title='ANALISIS')
    except:
        dst_sheet = dst_wb.create_sheet(title='ANALISIS')
    # Copy the contents of the source sheet to the destination sheet
    #for row in src_sheet.iter_rows(values_only=True):
    #    dst_sheet.append(row)
    data_sheet = dst_wb['DATA']
    copy_column(src_sheet, dst_sheet, 2, 1)
    corte_1(dst_sheet, 1, 2)
    corte_2(dst_sheet, 1, 2, 3)
    llave_res(dst_sheet, 1, 2, 3, 4)
    corte_3(dst_sheet, 1, 2, 5)
    local(dst_sheet, 5, 6)
    visita(dst_sheet, 5, 7)
    resul_final(dst_sheet,6,7,8)
    goleada(dst_sheet,6,7,9)

    resulx(data_sheet,106)
    cuenta(data_sheet,107)
    resr(data_sheet,108)
    coinci(data_sheet,110)
    choice_matchs(data_sheet, 100, 102, 105, 111,  0.5, 0.5, 'SI')
    # Assuming you have already loaded the workbook and selected the appropriate sheets
    #sheet_to_lookup = dst_wb['DATA']
    #lookup_sheet =dst_wb['ANALISIS']
    # Save the workbook
    # Save the changes to the destination Excel file
    # Set the value of a cell to a formula as text
    dst_sheet['M3'].value = '3'
    dst_sheet['M4'].value = '0'
    dst_sheet['M5'].value = '1'
    dst_sheet['N3'].value = '=COUNTIF(DATA!DF:DF,ANALISIS!M3)'
    dst_sheet['N4'].value = '=COUNTIF(DATA!DF:DF,ANALISIS!M4)'
    dst_sheet['N5'].value = '=COUNTIF(DATA!DF:DF,ANALISIS!M5)'
    dst_sheet['N6'].value = '=SUM(N4:N5)'
    dst_sheet['O3'].value = '=N3/$N$6'
    dst_sheet['O4'].value = '=N4/$N$6'
    dst_sheet['O5'].value = '=N5/$N$6'
    dst_sheet['L3'].value = '=M3&"|"&N3&"|"&O3'
    dst_sheet['L4'].value = '=M4&"|"&N4&"|"&O4'
    dst_sheet['L5'].value = '=M5&"|"&N5&"|"&O5'
    dst_sheet['M9'].value  = '3'
    dst_sheet['M10'].value = '0'
    dst_sheet['M11'].value = '1'
    dst_sheet['N9'].value  = '=COUNTIFS(DATA!DF:DF,ANALISIS!M9,DATA!DG:DG,"CHOICE")'
    dst_sheet['N10'].value = '=COUNTIFS(DATA!DF:DF,ANALISIS!M10,DATA!DG:DG,"CHOICE")'
    dst_sheet['N11'].value = '=COUNTIFS(DATA!DF:DF,ANALISIS!M11,DATA!DG:DG,"CHOICE")'
    dst_sheet['N12'].value = '=SUM(N10:N11)'
    dst_sheet['O9'].value  = '=N9/$N$12'
    dst_sheet['O10'].value = '=N10/$N$12'
    dst_sheet['O11'].value = '=N11/$N$12'
    dst_sheet['M14'].value  = '3'
    dst_sheet['M15'].value = '0'
    dst_sheet['M16'].value = '1'
    dst_sheet['N14'].value  = '=COUNTIFS(DATA!DF:DF,ANALISIS!M14,DATA!DH:DH,">=0.8")'
    dst_sheet['N15'].value = '=COUNTIFS(DATA!DF:DF,ANALISIS!M15,DATA!DH:DH,">=0.8")'
    dst_sheet['N16'].value = '=COUNTIFS(DATA!DF:DF,ANALISIS!M16,DATA!DH:DH,">=0.8")'
    dst_sheet['N17'].value = '=SUM(N15:N16)'
    dst_sheet['O14'].value  = '=N14/$N$17'
    dst_sheet['O15'].value = '=N15/$N$17'
    dst_sheet['O16'].value = '=N16/$N$17'
    dst_sheet['M19'].value  = '3'
    dst_sheet['M20'].value = '0'
    dst_sheet['M21'].value = '1'
    dst_sheet['N19'].value  = '=COUNTIFS(DATA!DF:DF,ANALISIS!M14,DATA!DI:DI,">=0.8")'
    dst_sheet['N20'].value = '=COUNTIFS(DATA!DF:DF,ANALISIS!M15,DATA!DI:DI,">=0.8")'
    dst_sheet['N21'].value = '=COUNTIFS(DATA!DF:DF,ANALISIS!M16,DATA!DI:DI,">=0.8")'
    dst_sheet['N17'].value = '=SUM(N20:N21)'
    dst_sheet['O19'].value  = '=N19/$N$22'
    dst_sheet['O20'].value = '=N20/$N$22'
    dst_sheet['O21'].value = '=N21/$N$22'

    dst_wb.save(pathb + file_nameb)
    src_wb.save(patha + file_name)
    dst_wb.close()
    src_wb.close()

    # Load the original workbook
    workbook = openpyxl.load_workbook(pathb + file_nameb)

    # Create a copy of the workbook
    copy_workbook = openpyxl.Workbook()

    # Copy the contents of the original workbook to the copy workbook
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        copy_worksheet = copy_workbook.create_sheet(sheet_name)
        for row in worksheet.iter_rows():
            copy_row = [cell.value for cell in row]
            copy_worksheet.append(copy_row)

    # Get the sheet you want to modify
    sheet = copy_workbook['DATA']

    #sheet.delete_cols(2, 103, shiftCols=False)

    # Loop through columns 2 to 4 and delete the contents of each cell
    for column in range(2, 103):
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=column)
            cell.value = None

    sheet_to_remove = copy_workbook['Sheet']
    copy_workbook.remove(sheet_to_remove)
    # Save the modified workbook
    copy_workbook.save(patha + file_namec)
    workbook.close()
    copy_workbook.close()

    time.sleep(segc)
    pyautogui.hotkey('win','r')
    pyautogui.write(pathb + file_nameb ,interval = 0.5)
    pyautogui.press('enter',interval = 0.5)
    time.sleep(segc)
    pyautogui.hotkey('ctrl','s',interval = 0.5)
    time.sleep(segc)
    pyautogui.hotkey('alt','f4',interval = 0.5)
    time.sleep(segc)
    time.sleep(segc)

    send_email_y(subject, message, to_email, from_email, password,pathb + file_nameb,file_nameb)
    print('Se envio el correo admin')
    send_email_y(subject, message, to_email, from_email, password,patha + file_namec,file_namec)
    send_email_y(subject, message, 'rtriguerosvargas@gmail.com', from_email, password,patha + file_namec,file_namec)
    send_email_y(subject, message, 'Gustavocarazasc@gmail.com', from_email, password,patha + file_namec,file_namec)
    send_email_y(subject, message, 'Willmezamotta@gmail.com', from_email, password,patha + file_namec,file_namec)
    send_email_y(subject, message, 'Ricardo.950219@gmail.com', from_email, password,patha + file_namec,file_namec)
    send_email_y(subject, message, 'a.escrib137@gmail.com', from_email, password,patha + file_namec,file_namec)
    send_email_y(subject, message, 'Zadihr90@gmail.com', from_email, password,patha + file_namec,file_namec)
    send_email_y(subject, message, 'mamoblitas@gmail.com', from_email, password,patha + file_namec,file_namec)
    send_email_y(subject, message, 'jhmciurliza@gmail.com', from_email, password,patha + file_namec,file_namec)
    print('Se envio el correo a los clientes')
    #SELENIUM ---------------------------XXXXXXXXXXXXXXXXX

    #SOUP ---------------------------XXXXXXXXXXXXXXXXX
except Exception as e:

    message = f"Error occurred: {str(e)}"
    trace = traceback.format_exc()
    x = f"{message}\n{trace}"
    send_email_x('Error', str(x)  , to_email, from_email, password)
    #logger.exception("Exception Occured while code Execution: "+ str(e))
    print(x)
    print('Se envio el correo de error')
